#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import django
import requests
import openpyxl
import datetime
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

CURRENT_PATH = os.path.realpath(os.path.dirname(__file__))
PROJECT_ROOT = os.path.realpath(os.path.join(CURRENT_PATH, os.pardir))
sys.path.insert(0, PROJECT_ROOT)

from base import load_django_settings

load_django_settings()
django.setup()

from django.conf import settings
from operation.models import Token
from operation.models import RoleInfo

APPID = settings.APPID
SECRET = settings.APP_SECRET


def import_army_data(excel_path):

    wb = load_workbook(excel_path)
    ws = wb.active

    fail_num = success_num = 0
    for i in range(1,ws.max_row+1):
        try:
            roleinfo = RoleInfo()
            for j in range(1,ws.max_column-1):
                roleinfo.id = ws.cell(row=i, column=1).value
                roleinfo.name = ws.cell(row=i, column=4).value
                roleinfo.abstract = ws.cell(row=i, column=5).value
                roleinfo.role_tv = ws.cell(row=i, column=6).value
                roleinfo.role_img_url = ws.cell(row=i, column=7).value
                roleinfo.share_ico_url = ws.cell(row=i, column=8).value
                roleinfo.create_time = datetime.datetime.now()
                roleinfo.modify_time_time = datetime.datetime.now()
            roleinfo.save()
        except:
            fail_num += 1
            continue
        success_num += 1
    print '导入完成，成功%s条，失败%s条'%(success_num,fail_num)

if __name__ == "__main__":
    excel_path = 'operation_roleinfo.xlsx'
    import_army_data(excel_path)
